import os
import sys
import argparse
from time import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from create_pickle import log_replace
sys.path.append(os.path.join(os.path.dirname(os.path.realpath(__file__)), "src"))
from beatsearch.data.rhythmcorpus import RhythmCorpus
from beatsearch.utils import print_progress_bar
from beatsearch.data.rhythm import TrackDistanceMeasure

SIM_MEASURES = TrackDistanceMeasure.get_measures()
SIM_MEASURE_NAMES = tuple(SIM_MEASURES.keys())
SIM_MEASURE_CLASSES = tuple(SIM_MEASURES.values())


CELL_COLORS = {
    'red': PatternFill(start_color=colors.RED, end_color=colors.RED, fill_type='solid'),
    'blue': PatternFill(start_color=colors.BLUE, end_color=colors.BLUE, fill_type='solid'),
    'green': PatternFill(start_color=colors.GREEN, end_color=colors.GREEN, fill_type='solid')
}


def get_args():
    parser = argparse.ArgumentParser(
        description="Creates a spreadsheet with rhythm distance "
                    "tables with one worksheet per similarity measure.")
    parser.add_argument("output", type=str,
                        help="The output *.xls file")
    parser.add_argument("--corpus", type=argparse.FileType('r'),
                        help="The *.pkl file containing the rhythms",
                        default="./data/rhythms.pkl")
    parser.add_argument("--track", type=int, default=36,
                        help="The rhythm track to compare")
    return parser.parse_args()


if __name__ == '__main__':
    args = get_args()

    # load corpus and get rhythm names for spreadsheet columns and rows headers
    log_replace("Loading rhythms from: %s" % args.corpus.name)
    corpus = RhythmCorpus.load(args.corpus)
    rhythm_names = [rhythm.name for rhythm in corpus]
    track_name = args.track
    log_replace("Loaded rhythms from '%s' containing %i rhythms\n" % (args.corpus.name, len(corpus)))

    # create spreadsheet with one worksheet per distance measure (truncate worksheet names on 31 chars)
    wb = Workbook(write_only=True)
    ws_gen = ((wb.create_sheet("%s.." % name[:29] if len(name) > 31 else name),
               tsm()) for name, tsm in SIM_MEASURES.items())

    # for progress bar and remaining time estimation
    n_steps, step_i, row_i, ws_i = len(corpus) * len(corpus) * len(SIM_MEASURES), 0, 0, 0
    t_start = time()

    def print_progress(worksheet, worksheet_i):
        print_progress_bar(
            step_i + 1, n_steps,
            "Computing track distances", "[%s (%i/%i), %i/%i]" % (
                str(worksheet.title), worksheet_i + 1, len(SIM_MEASURES), row_i + 1, len(corpus)
            ),
            fill='O', length=25, decimals=2, starting_time=t_start
        )

    # rhythms as column headers
    column_headers = [""] + rhythm_names

    for ws, measure in ws_gen:
        ws.append(column_headers)
        row_i = 0

        for rhythm_a in corpus:
            track_a = rhythm_a.get_track(args.track)

            if track_a is None:
                row_i += 1
                step_i += len(corpus)
                continue

            def create_row_generator():
                global step_i
                yield rhythm_names[row_i]  # row header

                for rhythm_b in corpus:
                    print_progress(ws, ws_i)
                    track_b = rhythm_b.get_track(args.track)
                    if track_b is None:
                        yield 'Track (%s) is None' % args.track
                    else:
                        try:
                            yield measure.get_distance(track_a, track_b)
                        except Exception as e:
                            yield e.__class__.__name__
                    step_i += 1

            ws.append(create_row_generator())
            row_i += 1

        ws_i += 1

    save_msg = "Saving spreadsheet to: %s" % args.output
    log_replace(save_msg)
    wb.save(args.output)
    log_replace("%s [Done]" % save_msg)

