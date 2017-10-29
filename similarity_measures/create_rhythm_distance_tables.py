import os
import sys
import argparse
from time import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from beatsearch.utils import print_progress_bar
from create_pickle import create_ascii_spinner, log_replace
from matplotlib import pyplot as plt
sys.path.append(os.path.join(os.path.dirname(os.path.realpath(__file__)), "src"))
from beatsearch.data.rhythm import Rhythm
from beatsearch.data.rhythmcorpus import RhythmCorpus


def parse_args():
    parser = argparse.ArgumentParser(
        description="Creates a spreadsheet with rhythm distance "
                    "tables with one worksheet per similarity measure.")
    parser.add_argument("corpus", type=argparse.FileType('r'),
                        help="The *.pkl file containing the rhythms")
    parser.add_argument("output", type=str,
                        help="The output *.xls file")
    parser.add_argument("--track", type=int, default=36,
                        help="The rhythm track to compare")
    return parser.parse_args()


red_fill = PatternFill(start_color=colors.RED, end_color=colors.RED, fill_type='solid')
blue_fill = PatternFill(start_color=colors.BLUE, end_color=colors.BLUE, fill_type='solid')
green_fill = PatternFill(start_color=colors.GREEN, end_color=colors.GREEN, fill_type='solid')


if __name__ == '__main__':
    args = parse_args()

    # load corpus and get rhythm names for spreadsheet columns and rows headers
    log_replace("Loading rhythms from: %s" % args.corpus.name)
    corpus = RhythmCorpus.load(args.corpus)
    rhythm_names = [""] + [rhythm.name for rhythm in corpus]
    track_name = args.track
    log_replace("Loaded rhythms from '%s' containing %i rhythms\n" % (args.corpus.name, len(corpus)))

    # create spreadsheet and remove the first
    wb = Workbook()
    wb.remove_sheet(wb.active)

    # create one worksheet per rhythm track distance measure
    worksheets = ((
            wb.create_sheet('Hamming distance'),
            Rhythm.Track.get_hamming_distance_to
        ), (
            wb.create_sheet('Euclidean IOI vector distance'),
            Rhythm.Track.get_euclidean_inter_onset_vector_distance_to
        )
    )

    # set column headers
    for ws, _ in worksheets:
        ws.append(rhythm_names)

    spinner = spinner = create_ascii_spinner("...ooOO@*         ")

    # begin iterating at second row (first row is for headers)
    row_i = 2
    computation_i = 1
    n_total_computations = len(corpus) * len(corpus)
    t_start = time()

    for rhythm_a in corpus:
        track_a = rhythm_a.get_track(track_name)

        # set row headers
        for ws, _ in worksheets:
            ws.cell(row=row_i, column=1, value=rhythm_a.name)

        # skip this rhythm if it doesn't have the desired track
        if track_a is None:
            row_i += 1
            computation_i += len(corpus)
            continue

        # begin iterating at second column (first column is for headers)
        col_i = 2

        for rhythm_b in corpus:
            print_progress_bar(computation_i, n_total_computations, "Computing distances",
                               "[%i/%i]" % (computation_i, n_total_computations), 2, starting_time=t_start)
            track_b = rhythm_b.get_track(track_name)

            # skip this rhythm if it doesn't have the desired track
            if track_b is None:
                col_i += 1
                computation_i += 1
                continue

            # compute and fill in the distances from track_a to track_b
            for ws, dist_measure in worksheets:
                fill = green_fill

                try:
                    dist = dist_measure.im_func(track_a, track_b)
                    if dist == 0:
                        fill = blue_fill
                except ValueError as e:
                    dist = e.message
                    fill = red_fill

                cell = ws.cell(row=row_i, column=col_i, value=dist)
                cell.fill = fill

            computation_i += 1
            col_i += 1
        row_i += 1

    save_msg = "Saving spreadsheet to: %s" % args.output
    log_replace(save_msg)
    wb.save(args.output)
    log_replace("%s [Done]" % save_msg)

